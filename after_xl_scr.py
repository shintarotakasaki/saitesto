# -*- coding: utf-8 -*-
"""after_xl_scr
Automatically generated by Colab.
Original file is located at
    https://colab.research.google.com/drive/1b3VXkHSUL8qaZE8LvrnmW17Nh3gh3TGN
"""

from openpyxl import load_workbook
import streamlit as st
from io import BytesIO
import requests
import shutil

# コメントアウトされたコード

def xl_data_upload():
    values = []
    
    #after_xl = st.file_uploader("アフター申請書エクセルをアップロードしてください")

    if after_xl is not None:
        file_mime = after_xl.type
        if file_mime == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
            try:
                file = BytesIO(after_xl.getvalue())
                wb = load_workbook(filename=file)
                sheet = wb.active
                #st.write(f"Sheet title: {sheet.title}")
                rects = ["G19","G20","L15","N20","G21","N21","G23","Q23","H28"]

                for rect in rects:
                #残しておきたいので取りあえずコメントアウト、あとで消す。
                    
                    #points,label = rect
                    #labels.append(label)
                    #for point in points:
                        #values.append(sheet[point].value)
                        
                    values.append(sheet[rect].value)

                for rect, value in zip(rects, values):
                    st.write(f"{rect}: {value}")
                
                return values

            except Exception as e:
                st.error(f"エラーが発生しました: {e}")
                return None, None

        else:
            st.write("エクセルファイル(.xlsx)をアップロードしてください")
            st.stop()

        #残しておきたいので取りあえずコメントアウト、あとで消す。
        """
        for label, value in zip(labels, values):
            st.write(f"**{label}**: {value}")
        """

def afterxl_dataget ():
    """
    GitHubからExcelファイルをダウンロードし、開く関数。
    """
    xlpoints = ["AC9-1","AC9","AC11","AC13","AC15","AC17","AC19","A11","S11","AM9"]
    
    after_xl = st.file_uploader("アフター申請書エクセルをアップロードしてください")
    
    if after_xl is not None:
        xl_data_upload(after_xl):
    
    else:
        st.write("エクセルファイル(.xlsx)をアップロードしてください")
        st.stop()

    
    try:
        github_url = "https://github.com/shintarotakasaki/excel3/raw/main/伝票(規格品)_ラベル_指示書.xlsm"
        # ファイルをダウンロードして一時ファイルとして保存
        response = requests.get(github_url,stream=True)
        if response.status_code == 200:
            file_path = "伝票(規格品)_ラベル_指示書.xlsm"  # ここで file_path を定義
            with open("伝票(規格品)_ラベル_指示書.xlsm",'wb')as f:
                response.raw.decode_content = True
                shutil.copyfileobj(response.raw, f)
    
            wb_demp = load_workbook(file_path, keep_vba=True)
            ws_demp = wb_demp['納品書控(製品)']
            wb_demp.active = ws_demp

            # ここでwb_dempを使って処理を行う
        else:
            st.error(f"ファイルのダウンロード中にエラーが発生しました: {response.status_code}")

    except Exception as e:  # 例外が発生した場合の処理
        st.error(f"エラーが発生しました: {e}")

if __name__ == "__main__":
    afterxl_dataget()
# ... (他のコードは省略)
